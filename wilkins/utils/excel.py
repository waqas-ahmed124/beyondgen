import pandas as pd
import os
import re
import zipfile
import openpyxl
from io import BytesIO
import base64
from collections import defaultdict
from PIL import Image, ExifTags
import hashlib
import logging

logger = logging.getLogger(__name__)


def dataframe_to_excel(df, excel_path):
    """
    Convert a DataFrame to an Excel file.
    """
    df.to_excel(excel_path, index=False)


def zip_file(file_path, zip_path):
    """
    Zip a file at the given path.
    """
    with zipfile.ZipFile(zip_path, "w") as zipf:
        zipf.write(file_path, os.path.basename(file_path))


def detect_sheet_and_header(file_path, probe_rows=5):
    """
    Detect the correct sheet and header row based on the given heuristic.
    """
    # Load the Excel file without any data to access the sheet names
    xls = pd.ExcelFile(file_path)

    if len(xls.sheet_names) == 1:
        rfp_sheets = [xls.sheet_names[0]]
    else:
        rfp_sheets = [sheet for sheet in xls.sheet_names]

    target_sheest = []
    target_header_rows = []

    for sheet in rfp_sheets:
        df_probe = pd.read_excel(
            file_path, sheet_name=sheet, nrows=probe_rows, header=None
        )

        # Checking each row for the presence of the specified values
        for idx, row in df_probe.iterrows():
            if any(val in row.values for val in ["Market", "Vendor", "Size"]):
                target_sheest.append(sheet)
                target_header_rows.append(idx)
                break

    return target_sheest, target_header_rows


def extract_images_to_df(file_path):
    """Extract the images from the .xlsx file and returns those images as bytes"""
    images_data = []

    # Open the .xlsx file as a zip file
    with zipfile.ZipFile(file_path, "r") as zip_ref:
        # Iterate through each file in the zip
        for file in zip_ref.namelist():
            # Adjusted to handle different image storage
            if file.startswith("xl/media/"):
                img_data = zip_ref.read(file)
                images_data.append(img_data)

    return images_data


def correct_image_orientation(img):
    """Correct orientation of an image based on its EXIF data"""
    try:
        exif = dict(img._getexif().items())
        for orientation in ExifTags.TAGS.keys():
            if ExifTags.TAGS[orientation] == "Orientation":
                if exif[orientation] == 3:
                    img = img.rotate(180, expand=True)
                elif exif[orientation] == 6:
                    img = img.rotate(270, expand=True)
                elif exif[orientation] == 8:
                    img = img.rotate(90, expand=True)
                break
    except (AttributeError, KeyError, IndexError, TypeError):
        # Cases where the image doesn't have EXIF data or orientation tag
        pass
    return img


def extract_all_images(sheet):
    """Extract all images from the sheet"""
    images = []
    for image in sheet._images:
        img = Image.open(image.ref)
        img = correct_image_orientation(img)  # Correct orientation
        images.append(img)
    return images


def has_required_headers(sheet, required_headers):
    """Check if the sheet has the required headers"""
    for row in sheet.iter_rows(values_only=True):
        if all(header in row for header in required_headers):
            return True
    return False


def extract_images_to_dict(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    images_dict = defaultdict(list)

    if "Photos" in wb.sheetnames:
        sheet = wb["Photos"]
        all_images = extract_all_images(sheet)

        # Identify label rows
        label_rows = [
            row
            for row in range(1, sheet.max_row + 1)
            if any(sheet.cell(row, col).value for col in range(1, sheet.max_column + 1))
        ]
        # Create labels based on row values
        labels = [sheet.cell(row, 1).value or f"Label_{row}" for row in label_rows]

        # Distribute images between label rows
        images_per_label_section = len(all_images) // len(label_rows)
        image_index = 0
        for i in range(len(label_rows)):
            label = labels[i]
            start_index = image_index
            end_index = start_index + images_per_label_section
            if i == len(label_rows) - 1:
                # Assign remaining images to the last label
                end_index = len(all_images)
            for img in all_images[start_index:end_index]:
                images_dict[label].append(img)
            image_index = end_index
    else:
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # Check if the sheet has the required headers
            required_headers = ["Market", "Vendor", "Size"]
            if has_required_headers(sheet, required_headers):
                all_images = extract_all_images(sheet)

                for img in all_images:
                    label = f"unlabeled"
                    images_dict[label].append(img)

    wb.close()
    return dict(images_dict)


def image_to_base64(image):
    # Convert the image to its base64-encoded representation
    image_stream = BytesIO()
    image_rgb = image.convert("RGB")  # Convert to RGB mode
    image_rgb.save(image_stream, format="JPEG")
    image_stream.seek(0)
    return base64.b64encode(image_stream.read()).decode("utf-8")


def remove_duplicates_from_all_images(all_images):
    # Create a new dictionary to store unique images for each label
    unique_images_dict = {}
    for label, images in all_images.items():
        unique_images = []
        seen_images = set()

        for image in images:
            # Convert the image to base64-encoded representation
            image_data = image_to_base64(image)

            if image_data not in seen_images:
                unique_images.append(image)
                seen_images.add(image_data)

        unique_images_dict[label] = unique_images

    return unique_images_dict


def append_images_to_df(file_path, target_sheets, target_header_rows, all_images):
    all_data = {}

    # Remove duplicates from the all_images dictionary
    unique_images_dict = remove_duplicates_from_all_images(all_images)

    for sheet_name, header_row in zip(target_sheets, target_header_rows):
        data_df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
        data_df = data_df.dropna(subset=["Market", "Vendor", "Size"])
        data_df["Images"] = None

        # Create a copy of unique_images_dict to avoid modifying the original dictionary
        available_images = {
            key: images.copy() for key, images in unique_images_dict.items()
        }

        for idx, row in data_df.iterrows():
            assigned_image = None

            for key in available_images.keys():
                if any(
                    key in cell.lower() or cell.lower() in key
                    for key in available_images.keys()
                    for cell in row.astype(str).str.lower()
                ):
                    if available_images[key]:
                        assigned_image = available_images[key].pop()
                        break  # Stop searching once an image is assigned

            if assigned_image is None:
                # If no images found in associated columns, use an image from "unlabeled" (if available)
                if "unlabeled" in available_images and available_images["unlabeled"]:
                    assigned_image = available_images["unlabeled"].pop()
                else:
                    assigned_image = None

            # Convert the assigned image to RGB mode if it's in RGBA mode
            if assigned_image and assigned_image.mode == "RGBA":
                assigned_image = assigned_image.convert("RGB")

            # Now you can save the image as JPEG
            if assigned_image:
                image_stream = BytesIO()  # Reset the image stream
                assigned_image.save(image_stream, format="JPEG")

                # Get the image data in bytes
                image_data = image_stream.getvalue()

                # Assign the image data to the DataFrame
                data_df.at[idx, "Images"] = image_data

        all_data[sheet_name] = data_df

    return all_data


def combine_dataframes(dfs):
    if not dfs:
        return None

    combined_df = dfs[0]
    for df in dfs[1:]:
        combined_df = combined_df.concat(
            df, on=["Vendor", "Unit #"], how="outer", suffixes=[None, None]
        )

    return combined_df


# Function to generate MD5 hash for each row
# We use this function to create a unique unit number for each row in a df
def generate_md5_as_number(row):
    # Check if the columns exist in the DataFrame
    media_type = (
        str(row.get("Media Type", "")) if not pd.isna(row.get("Media Type")) else ""
    )
    vendor = str(row.get("Vendor", "")) if not pd.isna(row.get("Vendor")) else ""
    town = str(row.get("Town ", "")) if not pd.isna(row.get("Town ")) else ""
    unit_number = str(row.get("Unit #", "")) if not pd.isna(row.get("Unit #")) else ""

    data_string = f"{media_type}_{vendor}_{town}_{unit_number}"
    return hashlib.md5(data_string.encode()).hexdigest()


def clean_up(df: pd.DataFrame):
    # Define mappings for replacements
    facing_replacements = {
        "North": "N",
        "South": "S",
        "East": "E",
        "West": "W",
        "Northeast": "NE",
        "Northwest": "NW",
        "Southeast": "SE",
        "Southwest": "SW",
        "South East": "SE",
        "South West": "SW",
        "North East": "NE",
        "North West": "NW",
        "South-East": "SE",
        "South-West": "SW",
        "North-East": "NE",
        "North-West": "NW",
    }
    illuminated_replacements = {"Yes": "Y", "No": "N"}
    # For the 'Facing' column if it exists
    if "Facing" in df.columns:
        df["Facing"] = df["Facing"].astype(str).str.title().replace(facing_replacements)

    # For the 'Illuminated?' column if it exists
    illuminated_column = next(
        (col for col in df.columns if "Illuminated?\n(Y or N)" in col), None
    )
    if illuminated_column:
        df[illuminated_column] = (
            df[illuminated_column]
            .astype(str)
            .str.capitalize()
            .replace(illuminated_replacements)
        )

    # For the 'A18+ Weekly Impressions' column if it exists
    impressions_column = next(
        (col for col in df.columns if "A18+ Weekly Impressions" in col), None
    )

    def safe_format(x):
        try:
            # Attempt to convert to float and format
            return f"{float(x):,.0f}"
        except (ValueError, TypeError):
            # If conversion fails, return the original value
            return x

    if impressions_column:
        df[impressions_column] = pd.to_numeric(
            df[impressions_column], errors="coerce"
        ).fillna(df[impressions_column])
        df[impressions_column] = df[impressions_column].apply(
            lambda x: safe_format(x) if pd.notnull(x) else x
        )

    def format_size(size):
        # Match patterns like "4.25'h x 3.16'w" or "(2) 4.25'h x 6.41'w"
        matches = list(
            re.finditer(r"\(?\s*(\d+)?\s*\)?\s*(\d+(\.\d+)?)'h x (\d+(\.\d+)?)'w", size)
        )

        # If no matches, return the original size string
        if not matches:
            return size

        formatted_sizes = []

        for match in matches:
            height = float(match.group(2))
            width = float(match.group(4))

            # Convert height to feet and inches
            height_feet = int(height)
            height_inches = round((height - height_feet) * 12)

            # Convert width to feet and inches
            width_feet = int(width)
            width_inches = round((width - width_feet) * 12)

            formatted_sizes.append(
                f"{height_feet}'{height_inches}\"h x {width_feet}'{width_inches}\"w"
            )

        return ", ".join(formatted_sizes)

    df["Size"] = df["Size"].apply(lambda x: format_size(str(x)) if pd.notnull(x) else x)
    # Format column with dollar sign
    rate_column = next((col for col in df.columns if "4 Week Media Rate " in col), None)
    if rate_column:
        df[rate_column] = df[rate_column].apply(
            lambda x: f"${float(x):,.2f}"
            if pd.notnull(x) and isinstance(x, (int, float))
            else (
                x
                if isinstance(x, str)
                and not x.lstrip("-").replace(".", "", 1).isdigit()
                else f"${float(x):,.2f}"
            )
        )
    rate_column = next((col for col in df.columns if "Production Cost" in col), None)
    if rate_column:
        df[rate_column] = df[rate_column].apply(
            lambda x: f"${float(x):,.2f}"
            if pd.notnull(x) and isinstance(x, (int, float))
            else (
                x
                if isinstance(x, str)
                and not x.lstrip("-").replace(".", "", 1).isdigit()
                else f"${float(x):,.2f}"
            )
        )
    rate_column = next((col for col in df.columns if "Installation Cost " in col), None)
    if rate_column:
        df[rate_column] = df[rate_column].apply(
            lambda x: f"${float(x):,.2f}"
            if pd.notnull(x) and isinstance(x, (int, float))
            else (
                x
                if isinstance(x, str)
                and not x.lstrip("-").replace(".", "", 1).isdigit()
                else f"${float(x):,.2f}"
            )
        )
    # Generate a unique identifier for each
    df["UniqueId"] = df.apply(generate_md5_as_number, axis=1)

    logger.info("Finished cleaning up the dataframe")
    return df
